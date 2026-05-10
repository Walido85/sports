"""
TuniWave Health Scraper — CNAM + SPOT
Python version to match existing tuniwave-scraper repo
Run: python health_scraper.py
"""

import requests
import json
import os
import csv
import time
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# ── CONFIG ────────────────────────────────────────────────────────────────────

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,ar-TN;q=0.8,en;q=0.6",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Cache-Control": "no-cache",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

GOVERNORATES = [
    "Ariana", "Ben Arous", "Bizerte", "Beja", "Jendouba",
    "Kairouan", "Mahdia", "Medenine", "Monastir", "Nabeul",
    "La Manouba", "Tunis", "Zaghouan", "Siliana", "Le Kef",
    "Sousse", "Kasserine", "Kebili", "Sidi Bouzid", "Sfax",
    "Gabes", "Gafsa", "Tozeur", "Tataouine",
]

# ── HTTP ──────────────────────────────────────────────────────────────────────

def get(url, params=None, timeout=20, retries=3):
    for attempt in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            return r
        except requests.exceptions.RequestException as e:
            if attempt == retries - 1:
                raise
            time.sleep(1 * (attempt + 1))

def get_html(url, params=None):
    r = get(url, params=params)
    r.encoding = r.apparent_encoding or "utf-8"
    return r.text

def get_binary(url):
    r = get(url)
    return r.content

# ── SAVE ──────────────────────────────────────────────────────────────────────

def save_json(name, data):
    path = os.path.join(OUTPUT_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    count = len(data) if isinstance(data, list) else 1
    print(f"💾 {name} — {count} records")

def save_csv(name, rows, columns):
    if not rows:
        return
    path = os.path.join(OUTPUT_DIR, name)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"📄 {name} — {len(rows)} rows")

# ── CNAM ──────────────────────────────────────────────────────────────────────

CNAM_BASE = "https://catalog.data.gov.tn"
CNAM_API  = f"{CNAM_BASE}/api/3/action/package_show?id=liste-des-prestataires-de-soins-conventionnes"
CNAM_PAGE = f"{CNAM_BASE}/fr/dataset/liste-des-prestataires-de-soins-conventionnes"

def cnam_get_download_urls():
    print("\n📡 CNAM — finding dataset files...")

    # 1. CKAN JSON API
    try:
        r = SESSION.get(CNAM_API, timeout=15, headers={**HEADERS, "Accept": "application/json"})
        r.raise_for_status()
        resources = r.json().get("result", {}).get("resources", [])
        urls = [
            {"url": res["url"], "format": res.get("format", "").upper(), "name": res.get("name", "")}
            for res in resources
            if any(ext in (res.get("format","") + res.get("url","")).lower() for ext in ["xlsx","xls","csv","json","ods"])
        ]
        if urls:
            print(f"✅ CKAN API: {len(urls)} file(s) found")
            return urls
    except Exception as e:
        print(f"⚠️  CKAN API: {e}")

    # 2. HTML scrape
    try:
        print("   Trying HTML scrape...")
        html = get_html(CNAM_PAGE)
        soup = BeautifulSoup(html, "lxml")
        urls = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if any(href.lower().endswith(ext) for ext in [".xlsx", ".xls", ".csv", ".json", ".ods"]):
                full = href if href.startswith("http") else CNAM_BASE + href
                if not any(u["url"] == full for u in urls):
                    urls.append({"url": full, "format": href.rsplit(".", 1)[-1].upper(), "name": a.text.strip()})
        if urls:
            print(f"   HTML scrape: {len(urls)} file(s) found")
            return urls
    except Exception as e:
        print(f"⚠️  HTML scrape: {e}")

    print("❌ CNAM: no files found")
    return []

def parse_xlsx_bytes(content):
    if not HAS_OPENPYXL:
        print("⚠️  openpyxl not installed, skipping XLSX parse")
        return []
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            record = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            rows.append(record)
    return rows

def parse_csv_bytes(content):
    decoded = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(decoded.splitlines())
    return [dict(r) for r in reader]

def normalize_provider(raw):
    def get(*keys):
        for k in keys:
            for rk, rv in raw.items():
                if k.lower() in rk.lower():
                    return str(rv).strip()
        return ""

    phone = get("téléphone", "telephone", "phone", "tel", "gsm").replace(" ", "").replace(".", "")

    return {
        "nom":             get("nom"),
        "prenom":          get("prénom", "prenom"),
        "type":            get("type_prestataire", "type", "catégorie", "category"),
        "specialite":      get("spécialité", "specialite", "specialty"),
        "gouvernorat":     get("gouvernorat", "governorate", "région"),
        "delegation":      get("délégation", "delegation"),
        "adresse":         get("adresse", "address"),
        "telephone":       phone,
        "convention_cnam": "oui",
        "type_garde":      "",
        "source":          "CNAM",
    }

def download_and_parse(file_info):
    try:
        print(f"   ⬇️  {file_info['name'] or file_info['url']}")
        content = get_binary(file_info["url"])
        fmt = file_info["format"].lower()

        if "csv" in fmt or file_info["url"].endswith(".csv"):
            rows = parse_csv_bytes(content)
        elif any(x in fmt for x in ["xls", "xlsx", "ods"]):
            rows = parse_xlsx_bytes(content)
        elif "json" in fmt:
            data = json.loads(content)
            rows = data if isinstance(data, list) else data.get("records", data.get("data", []))
        else:
            rows = parse_xlsx_bytes(content)  # try xlsx by default

        print(f"   ✅ {len(rows)} rows parsed")
        return rows
    except Exception as e:
        print(f"   ❌ Failed: {e}")
        return []

def scrape_cnam():
    urls = cnam_get_download_urls()
    if not urls:
        return []

    all_rows = []
    # Download concurrently
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(download_and_parse, u): u for u in urls}
        for future in as_completed(futures):
            rows = future.result()
            all_rows.extend(rows)

    normalized = [normalize_provider(r) for r in all_rows]
    return [r for r in normalized if r["nom"] or r["telephone"]]

# ── SPOT ──────────────────────────────────────────────────────────────────────

SPOT_BASE = "https://spot.tn"

def get_delegations(ville):
    try:
        html = get_html(f"{SPOT_BASE}/", params={"mayor": "pharmacie", "ville": ville})
        soup = BeautifulSoup(html, "lxml")
        select = soup.find("select", {"name": lambda n: n and "deleg" in n.lower()})
        if not select:
            return [{"value": "", "label": "Toutes"}]
        options = []
        for opt in select.find_all("option"):
            v = opt.get("value", "")
            t = opt.text.strip()
            if v and v not in ("0", ""):
                options.append({"value": v, "label": t})
        return options if options else [{"value": "", "label": "Toutes"}]
    except:
        return [{"value": "", "label": "Toutes"}]

def parse_pharmacies(html, ville, delegation_label, type_garde):
    soup = BeautifulSoup(html, "lxml")
    results = []

    # Table rows
    for row in soup.select("table tbody tr"):
        cols = [td.text.strip() for td in row.find_all("td")]
        if len(cols) >= 2 and cols[0]:
            results.append({
                "nom":         cols[0],
                "adresse":     cols[1] if len(cols) > 1 else "",
                "telephone":   cols[2].replace(" ", "") if len(cols) > 2 else "",
                "horaires":    cols[3] if len(cols) > 3 else "",
                "gouvernorat": ville,
                "delegation":  delegation_label,
                "type_garde":  type_garde,
                "specialite":  "",
                "prenom":      "",
                "type":        "Pharmacie",
                "convention_cnam": "",
                "source":      "SPOT",
            })

    # Card/div fallback
    if not results:
        for card in soup.select(".pharmacie, .pharmacy, [class*='pharm'], .result-item"):
            nom = card.find(["h3", "h4", "strong"])
            tel_tag = card.find("a", href=lambda h: h and h.startswith("tel:"))
            nom_text = nom.text.strip() if nom else ""
            tel_text = (tel_tag["href"].replace("tel:", "") if tel_tag else "").replace(" ", "")
            if nom_text:
                results.append({
                    "nom": nom_text,
                    "adresse": "",
                    "telephone": tel_text,
                    "horaires": "",
                    "gouvernorat": ville,
                    "delegation": delegation_label,
                    "type_garde": type_garde,
                    "specialite": "",
                    "prenom": "",
                    "type": "Pharmacie",
                    "convention_cnam": "",
                    "source": "SPOT",
                })

    return results

def search_spot(ville, delegation, type_garde):
    params = {
        "mayor": "pharmacie",
        "ville": ville,
        "delegation": delegation["value"],
        "type": "nuit" if type_garde == "Nuit" else "jour",
    }
    try:
        html = get_html(f"{SPOT_BASE}/", params=params)
        return parse_pharmacies(html, ville, delegation["label"], type_garde)
    except:
        return []

def scrape_spot():
    print(f"\n📡 SPOT — fetching pharmacies across {len(GOVERNORATES)} governorates...")

    # Build all tasks: city × delegation × day/night
    tasks = []
    print("   Getting delegations...")
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(get_delegations, v): v for v in GOVERNORATES}
        city_delegations = {}
        for future in as_completed(futures):
            ville = futures[future]
            city_delegations[ville] = future.result()

    for ville in GOVERNORATES:
        for delegation in city_delegations[ville]:
            for type_garde in ["Jour", "Nuit"]:
                tasks.append((ville, delegation, type_garde))

    print(f"   {len(tasks)} search tasks queued")

    all_results = []
    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = {ex.submit(search_spot, t[0], t[1], t[2]): t for t in tasks}
        for future in as_completed(futures):
            results = future.result()
            if results:
                t = futures[future]
                print(f"   ✅ {t[0]}/{t[1]['label']}/{t[2]}: {len(results)}")
                all_results.extend(results)

    return all_results

# ── DEDUP + MERGE ─────────────────────────────────────────────────────────────

def dedup(records):
    seen_phone = set()
    seen_name  = set()
    unique = []
    for r in records:
        phone = r.get("telephone", "").replace(" ", "").replace(".", "")
        name  = f"{r.get('nom','')}|{r.get('gouvernorat','')}".lower()

        if phone and len(phone) >= 8:
            if phone in seen_phone:
                continue
            seen_phone.add(phone)
        else:
            if name in seen_name:
                continue
            seen_name.add(name)

        unique.append(r)
    return unique

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("╔═══════════════════════════════════════╗")
    print("║   TuniWave Health Scraper             ║")
    print("║   CNAM + SPOT | Python | Concurrent   ║")
    print("╚═══════════════════════════════════════╝")

    start = time.time()

    # Run both scrapers (CNAM then SPOT — SPOT is slow so give it full thread pool)
    cnam_raw = scrape_cnam()
    spot_raw = scrape_spot()

    print(f"\n📊 Raw: CNAM={len(cnam_raw)} | SPOT={len(spot_raw)}")

    merged = dedup(cnam_raw + spot_raw)

    COLUMNS = ["nom","prenom","type","specialite","gouvernorat","delegation",
               "adresse","telephone","convention_cnam","type_garde","source"]

    pharmacies  = [r for r in merged if "pharm" in (r["type"]+r["source"]).lower() or r["source"] == "SPOT"]
    doctors     = [r for r in merged if any(x in (r["type"]+r["specialite"]).lower() for x in ["médecin","doctor","spécial","chirurg"])]
    labs        = [r for r in merged if "labor" in r["type"].lower() or "biol" in r["type"].lower()]
    garde_nuit  = [r for r in merged if r.get("type_garde") == "Nuit"]

    elapsed = round(time.time() - start, 1)
    print(f"\n✅ Done in {elapsed}s")
    print(f"   Total unique:  {len(merged)}")
    print(f"   Pharmacies:    {len(pharmacies)}")
    print(f"   Doctors:       {len(doctors)}")
    print(f"   Labs:          {len(labs)}")
    print(f"   Garde nuit:    {len(garde_nuit)}")

    save_json("all_providers.json", merged)
    save_json("pharmacies.json", pharmacies)
    save_json("doctors.json", doctors)
    save_json("labs.json", labs)
    save_json("garde_nuit.json", garde_nuit)
    save_csv("all_providers.csv", merged, COLUMNS)
    save_csv("pharmacies.csv", pharmacies, COLUMNS)
    save_csv("doctors.csv", doctors, COLUMNS)

    save_json("summary.json", {
        "scraped_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "duration_seconds": elapsed,
        "totals": {
            "all": len(merged),
            "pharmacies": len(pharmacies),
            "doctors": len(doctors),
            "labs": len(labs),
            "garde_nuit": len(garde_nuit),
        }
    })

    print("\n📁 Files saved in ./output/")

if __name__ == "__main__":
    main()
